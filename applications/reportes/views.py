from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, DecimalField, ExpressionWrapper, Q
from applications.productos.models import Local,StockLocal
from applications.ventas.models import DetalleVenta
from applications.users.models import User
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

@login_required
def reporte_ventas_por_local(request):
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    local_id = request.GET.get('local')
    user_id = request.GET.get('usuario')
    filtros = Q()
    if fecha_desde:
        filtros &= Q(venta__fecha__date__gte=fecha_desde)
    if fecha_hasta:
        filtros &= Q(venta__fecha__date__lte=fecha_hasta)
    if local_id:
        filtros &= Q(venta__local__id=local_id)
    if user_id:
        filtros &= Q(venta__user_made__id=user_id)

    resultados = DetalleVenta.objects.filter(filtros).annotate(
        subtotal_calculado=ExpressionWrapper(
            F('cantidad') * F('precio_unitario'),
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )
    ).select_related('venta', 'producto', 'user_made', 'venta__local')

    total = resultados.aggregate(suma=Sum('subtotal_calculado'))['suma'] or 0

    if request.GET.get('export') == 'excel':
        # Crear un libro de trabajo y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas por Local"

        # Estilos
        header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        header_font = Font(bold=True)
        border = Border(left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        top=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        total_fill = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")
        currency_format = '#,##0.00'

        # Escribir los encabezados con estilo
        headers = ['Producto', 'Descripción', 'Precio Unitario', 'Cantidad', 'Subtotal', 'Fecha', 'Usuario', 'Local']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border

        # Escribir los datos con estilo
        for row_num, r in enumerate(resultados, 2):
            ws.cell(row=row_num, column=1, value=r.producto.nombre).alignment = align_left
            ws.cell(row=row_num, column=2, value=r.producto.descripcion).alignment = align_left  # Agregar descripción
            ws.cell(row=row_num, column=3, value=r.precio_unitario).number_format = currency_format
            ws.cell(row=row_num, column=4, value=r.cantidad)
            ws.cell(row=row_num, column=5, value=r.subtotal_calculado).number_format = currency_format
            ws.cell(row=row_num, column=6, value=r.venta.fecha.strftime('%d/%m/%Y %H:%M')).alignment = align_center
            ws.cell(row=row_num, column=7, value=r.user_made.username).alignment = align_left
            ws.cell(row=row_num, column=8, value=r.venta.local.nombre).alignment = align_left

            # Aplicar bordes a las celdas de datos
            for col_num in range(1, 9):
                ws.cell(row=row_num, column=col_num).border = border

        # Escribir el total con estilo
        total_row = len(resultados) + 2
        ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=total_row, column=1).alignment = align_left
        ws.cell(row=total_row, column=2, value="").alignment = align_center
        ws.cell(row=total_row, column=3, value="").alignment = align_center
        ws.cell(row=total_row, column=4, value="").alignment = align_center
        ws.cell(row=total_row, column=5, value=total).number_format = currency_format
        ws.cell(row=total_row, column=5).fill = total_fill
        ws.cell(row=total_row, column=5).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value="").alignment = align_center
        ws.cell(row=total_row, column=7, value="").alignment = align_center
        ws.cell(row=total_row, column=8, value="").alignment = align_center

        # Ajustar el ancho de las columnas
        column_widths = {
            1: 20,  # Producto
            2: 30,  # Descripción
            3: 15,  # Precio Unitario
            4: 10,  # Cantidad
            5: 15,  # Subtotal
            6: 20,  # Fecha
            7: 15,  # Usuario
            8: 15,  # Local
        }
        for col_num, width in column_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # Configurar la respuesta para descargar el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_ventas_por_local_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        # Guardar el archivo en la respuesta
        wb.save(response)
        return response

    # Enviar los resultados al template si no se solicita exportar a Excel
    context = {
        'locales': Local.objects.all(),
        'resultados': resultados,
        'total': total,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'usuarios':User.objects.all(),
        'local_id': local_id,
    }
    return render(request, 'reportes/reporte_ventas_por_local.html', context)

@login_required
def reporte_stock_por_local(request):
    local_id = request.GET.get('local')

    filtros = Q(cantidad__gt=0)  # Cambié 'stock' por 'cantidad'
    if local_id:
        filtros &= Q(local__id=local_id)

    resultados = StockLocal.objects.filter(filtros).select_related('producto', 'local')

    if request.GET.get('export') == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock por Local"

        # Estilos
        header_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        header_font = Font(bold=True)
        border = Border(left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        top=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        # Encabezados
        headers = ['ID Producto', 'Nombre', 'Descripción', 'Local', 'Stock']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border

        # Datos
        for row_num, r in enumerate(resultados, 2):
            ws.cell(row=row_num, column=1, value=r.producto.id).alignment = align_center
            ws.cell(row=row_num, column=2, value=r.producto.nombre).alignment = align_left
            ws.cell(row=row_num, column=3, value=r.producto.descripcion or '').alignment = align_left
            ws.cell(row=row_num, column=4, value=r.local.nombre).alignment = align_left
            ws.cell(row=row_num, column=5, value=r.cantidad).alignment = align_center

            for col_num in range(1, 6):
                ws.cell(row=row_num, column=col_num).border = border

        # Ancho de columnas
        widths = {1: 12, 2: 25, 3: 40, 4: 20, 5: 10}
        for col_num, width in widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="reporte_stock_por_local_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        wb.save(response)
        return response

    context = {
        'locales': Local.objects.all(),
        'resultados': resultados,
        'local_id': local_id,
    }
    return render(request, 'reportes/reporte_stock_por_local.html', context)